"""Excel rate/reserve exhibits (openpyxl).

One workbook per run containing:
  * Triangle_<line>  -- cumulative triangle + age-to-age factors + CDFs
  * IBNR_Summary     -- per line/origin: latest, CDF, CL ultimate/IBNR,
                        BF ultimate/IBNR
  * Trend            -- monthly completed totals + measured annualized trend
  * Rate_Development -- line-by-line six-step rate buildup
  * Credibility      -- limited-fluctuation credibility table

Numeric cells are written at a canonical exhibit precision -- dollar values
rounded to the cent (``DOLLAR_DECIMALS``), factors/ratios to 8 decimals
(``FACTOR_DECIMALS``) -- because openpyxl serializes floats via ``%.16g``,
which cannot round-trip full 17-significant-digit doubles.  At this canonical
precision the write->reopen round-trip is float-EXACT: tests reopen the
workbook and assert cell == the canonically rounded engine value with ``==``
(no tolerance).  Use :func:`dollars` / :func:`factor` to reproduce the
canonical values from engine output.

HONESTY TAG: exhibits in this repo are built from SYNTHETIC seeded data.
"""

from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font

from .pricing import RateDevelopment
from .reserving import ReservingResult

_BOLD = Font(bold=True)

DOLLAR_DECIMALS = 2   # dollar-valued cells: to the cent
FACTOR_DECIMALS = 8   # factors, CDFs, trends, credibility Z, ratios


def dollars(v) -> float:
    """Canonical exhibit value for a dollar amount (or count)."""
    return round(float(v), DOLLAR_DECIMALS)


def factor(v) -> float:
    """Canonical exhibit value for a factor/ratio/rate."""
    return round(float(v), FACTOR_DECIMALS)


def _header(ws, row, values):
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = _BOLD


def write_exhibits(path: str,
                   triangles: Dict[str, "object"],
                   cl: Dict[str, ReservingResult],
                   bf: Dict[str, ReservingResult],
                   trend: Dict[str, dict],
                   rate_dev: Dict[str, RateDevelopment],
                   credibility_rows) -> None:
    """Write the full exhibit workbook.

    Parameters
    ----------
    triangles : {line: 2D ndarray} cumulative runoff triangles (NaN = future)
    cl, bf : {line: ReservingResult}
    trend : {line: {"monthly_totals": list, "measured_annual_trend": float}}
    rate_dev : {line: RateDevelopment}
    credibility_rows : iterable of (label, n, full_credibility_n, z)
    """
    wb = Workbook()
    wb.remove(wb.active)

    for line, tri in triangles.items():
        ws = wb.create_sheet(f"Triangle_{line}")
        ws.cell(row=1, column=1, value=f"Cumulative paid triangle - {line} "
                                       "(SYNTHETIC seeded data)").font = _BOLD
        n_o, n_d = tri.shape
        _header(ws, 2, ["origin \\ dev"] + list(range(n_d)))
        for i in range(n_o):
            ws.cell(row=3 + i, column=1, value=i)
            for j in range(n_d):
                v = tri[i, j]
                if v == v:  # not NaN
                    ws.cell(row=3 + i, column=2 + j, value=dollars(v))
        r = 3 + n_o
        res = cl[line]
        ws.cell(row=r, column=1, value="ATA factor").font = _BOLD
        for j, f in enumerate(res.factors):
            ws.cell(row=r, column=2 + j, value=factor(f))
        ws.cell(row=r + 1, column=1, value="CDF").font = _BOLD
        for j, c in enumerate(res.cdf):
            ws.cell(row=r + 1, column=2 + j, value=factor(c))

    ws = wb.create_sheet("IBNR_Summary")
    _header(ws, 1, ["line", "origin", "latest_paid", "age", "cdf",
                    "cl_ultimate", "cl_ibnr", "bf_ultimate", "bf_ibnr"])
    r = 2
    for line in triangles:
        c, b = cl[line], bf[line]
        for i in range(len(c.latest)):
            ws.cell(row=r, column=1, value=line)
            ws.cell(row=r, column=2, value=i)
            ws.cell(row=r, column=3, value=dollars(c.latest[i]))
            ws.cell(row=r, column=4, value=int(c.age[i]))
            ws.cell(row=r, column=5, value=factor(c.cdf[c.age[i]]))
            ws.cell(row=r, column=6, value=dollars(c.ultimate[i]))
            ws.cell(row=r, column=7, value=dollars(c.ibnr[i]))
            ws.cell(row=r, column=8, value=dollars(b.ultimate[i]))
            ws.cell(row=r, column=9, value=dollars(b.ibnr[i]))
            r += 1
        ws.cell(row=r, column=1, value=f"{line} TOTAL").font = _BOLD
        ws.cell(row=r, column=6, value=dollars(c.total_ultimate))
        ws.cell(row=r, column=7, value=dollars(c.total_ibnr))
        ws.cell(row=r, column=8, value=dollars(b.total_ultimate))
        ws.cell(row=r, column=9, value=dollars(b.total_ibnr))
        r += 2

    ws = wb.create_sheet("Trend")
    _header(ws, 1, ["line", "measured_annual_trend"])
    r = 2
    for line, t in trend.items():
        ws.cell(row=r, column=1, value=line)
        ws.cell(row=r, column=2, value=factor(t["measured_annual_trend"]))
        for j, v in enumerate(t["monthly_totals"]):
            ws.cell(row=r, column=3 + j, value=dollars(v))
        r += 1

    ws = wb.create_sheet("Rate_Development")
    fields = [("completed_claims", dollars), ("member_months", dollars),
              ("experience_pmpm", dollars), ("annual_trend", factor),
              ("trend_months", dollars), ("trend_factor", factor),
              ("projected_pmpm", dollars), ("credibility_n", dollars),
              ("full_credibility_n", dollars), ("z", factor),
              ("manual_pmpm", dollars), ("blended_pmpm", dollars),
              ("target_loss_ratio", factor), ("required_pmpm", dollars)]
    _header(ws, 1, ["step"] + list(rate_dev.keys()))
    for i, (f, canon) in enumerate(fields):
        ws.cell(row=2 + i, column=1, value=f).font = _BOLD
        for j, (line, rd) in enumerate(rate_dev.items()):
            ws.cell(row=2 + i, column=2 + j, value=canon(getattr(rd, f)))

    ws = wb.create_sheet("Credibility")
    _header(ws, 1, ["case", "n", "full_credibility_n", "z"])
    for i, (label, n, full_n, z) in enumerate(credibility_rows):
        ws.cell(row=2 + i, column=1, value=label)
        ws.cell(row=2 + i, column=2, value=dollars(n))
        ws.cell(row=2 + i, column=3, value=dollars(full_n))
        ws.cell(row=2 + i, column=4, value=factor(z))

    wb.save(path)
