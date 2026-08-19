"""Excel round-trip: reopen the workbook and verify cells equal the
canonically rounded engine output EXACTLY (``==``, no tolerance).

Canonical exhibit precision (see ratecraft.exhibits): dollars to the cent,
factors/ratios to 8 decimals -- required because openpyxl serializes floats
with %.16g, which cannot round-trip 17-significant-digit doubles."""

import numpy as np
import pytest
from openpyxl import load_workbook

from eval.make_exhibits import LINES, build_run, make
from ratecraft.exhibits import dollars, factor


@pytest.fixture(scope="module")
def run():
    return build_run(seed=42)


@pytest.fixture(scope="module")
def wb_path(tmp_path_factory, run):
    p = tmp_path_factory.mktemp("exhibits") / "sample.xlsx"
    make(str(p), seed=42)
    return str(p)


@pytest.fixture(scope="module")
def wb(wb_path):
    return load_workbook(wb_path)


def test_expected_sheets(wb):
    expected = [f"Triangle_{ln}" for ln in LINES] + [
        "IBNR_Summary", "Trend", "Rate_Development", "Credibility"]
    assert wb.sheetnames == expected


def test_triangle_cells_round_trip_exact(wb, run):
    triangles, cl, _, _, _, _ = run
    for line in LINES:
        ws = wb[f"Triangle_{line}"]
        tri = triangles[line]
        n_o, n_d = tri.shape
        for i in range(0, n_o, 7):
            for j in range(n_d):
                v = tri[i, j]
                cell = ws.cell(row=3 + i, column=2 + j).value
                if np.isnan(v):
                    assert cell is None
                else:
                    assert cell == dollars(v)  # exact float equality


def test_factors_and_cdfs_round_trip_exact(wb, run):
    triangles, cl, _, _, _, _ = run
    for line in LINES:
        ws = wb[f"Triangle_{line}"]
        n_o = triangles[line].shape[0]
        res = cl[line]
        for j, f in enumerate(res.factors):
            assert ws.cell(row=3 + n_o, column=2 + j).value == factor(f)
        for j, c in enumerate(res.cdf):
            assert ws.cell(row=4 + n_o, column=2 + j).value == factor(c)


def test_ibnr_summary_round_trip_exact(wb, run):
    triangles, cl, bf, _, _, _ = run
    ws = wb["IBNR_Summary"]
    n_o = triangles[LINES[0]].shape[0]
    # first line block starts at row 2
    c, b = cl[LINES[0]], bf[LINES[0]]
    for i in range(0, n_o, 5):
        r = 2 + i
        assert ws.cell(row=r, column=1).value == LINES[0]
        assert ws.cell(row=r, column=3).value == dollars(c.latest[i])
        assert ws.cell(row=r, column=6).value == dollars(c.ultimate[i])
        assert ws.cell(row=r, column=7).value == dollars(c.ibnr[i])
        assert ws.cell(row=r, column=8).value == dollars(b.ultimate[i])
    total_row = 2 + n_o
    assert ws.cell(row=total_row, column=6).value == dollars(c.total_ultimate)
    assert ws.cell(row=total_row, column=9).value == dollars(b.total_ibnr)


def test_trend_sheet_round_trip_exact(wb, run):
    _, _, _, trend, _, _ = run
    ws = wb["Trend"]
    for r, line in enumerate(LINES, start=2):
        assert ws.cell(row=r, column=1).value == line
        assert (ws.cell(row=r, column=2).value
                == factor(trend[line]["measured_annual_trend"]))
        for j, v in enumerate(trend[line]["monthly_totals"]):
            assert ws.cell(row=r, column=3 + j).value == dollars(v)


def test_rate_development_round_trip_exact(wb, run):
    _, _, _, _, rate_dev, _ = run
    ws = wb["Rate_Development"]
    header = [ws.cell(row=1, column=2 + j).value for j in range(len(LINES))]
    assert header == list(LINES)
    rows = {ws.cell(row=r, column=1).value: r for r in range(2, 16)}
    canon = {"experience_pmpm": dollars, "trend_factor": factor,
             "projected_pmpm": dollars, "z": factor,
             "blended_pmpm": dollars, "required_pmpm": dollars}
    for j, line in enumerate(LINES):
        rd = rate_dev[line]
        for field, fn in canon.items():
            assert (ws.cell(row=rows[field], column=2 + j).value
                    == fn(getattr(rd, field)))


def test_credibility_table_round_trip_exact(wb, run):
    _, _, _, _, _, cred_rows = run
    ws = wb["Credibility"]
    for i, (label, n, full_n, z) in enumerate(cred_rows):
        assert ws.cell(row=2 + i, column=1).value == label
        assert ws.cell(row=2 + i, column=2).value == dollars(n)
        assert ws.cell(row=2 + i, column=3).value == dollars(full_n)
        assert ws.cell(row=2 + i, column=4).value == factor(z)


def test_workbook_deterministic(tmp_path, wb_path):
    import hashlib
    import zipfile

    def content_hash(path):
        h = hashlib.sha256()
        with zipfile.ZipFile(path) as z:
            for member in sorted(m for m in z.namelist()
                                 if "worksheets" in m):
                h.update(z.read(member))
        return h.hexdigest()

    p2 = tmp_path / "again.xlsx"
    make(str(p2), seed=42)
    assert content_hash(wb_path) == content_hash(str(p2))
